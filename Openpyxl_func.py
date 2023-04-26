import openpyxl
import pandas
import re

letter = [i for i in 'ABCDEFGHIJKLMN']
wbook = openpyxl.open('C:\\Users\\Дмитрий\\PycharmProjects\\Da\\documents\\Расписание уроков.xlsx')
ws = wbook.active


def WithoutTeacher(full_name, next_day):
    teachers = []
    lessons = []
    for sheet_teacher in ws['D49':'D75']:
        for cell_teacher in sheet_teacher:
            teachers.append(cell_teacher.value)
    for sheet_day in ws['B2':'B38']:
        for cell_day in sheet_day:
            if cell_day.value == f'{next_day}':
                position_day_1 = letter[cell_day.column + 1] + str(cell_day.row + 1)
                position_day_2 = letter[cell_day.column + 8] + str(cell_day.row + 8)
                position_day_3 = letter[cell_day.column + 9] + str(cell_day.row + 1)
                position_day_4 = letter[cell_day.column + 11] + str(cell_day.row + 8)
    for teacher in teachers:
        if re.search(f'{full_name}', f'{teacher}'):
            for sheet_cabinet in ws['D49':'D75']:
                for cell_cabinet in sheet_cabinet:
                    if cell_cabinet.value == teacher:
                        cabinet = letter[cell_cabinet.column] + str(cell_cabinet.row)
                        cabinet_number = ws[f'{cabinet}']
                        for sheet in ws[f'{position_day_1}':f'{position_day_2}']:
                            for cell in sheet:
                                if re.search(cabinet_number.value[1:4], f'{cell.value}'):
                                    pos_les = letter[cell.column - 1] + str(cell.row)
                                    lessons.append([cell.column, cell.value])
                                    ws[f'{pos_les}'] = ''
    for sheet_lesson in ws[f'{position_day_1}':f'{position_day_2}']:
        for cell_lesson in sheet_lesson:
            if cell_lesson.value == '':
                ln = []
                k = 0
                pln_1 = letter[cell_lesson.column - 1] + str(cell_lesson.row)
                pln_2 = letter[cell_lesson.column - 1] + str(cell_lesson.row + 7)
                for sln in ws[f'{pln_1}':f'{pln_2}']:
                    if k < 1:
                        for cln in sln:
                            if cln.value == '' or cln.value is None:
                                pass
                            elif re.fullmatch(r'[1]?[019].', f'{cln.value}'):
                                k += 1
                                break
                            else:
                                ln.append(cln.value)
                                ws[f'{letter[cln.column - 1] + str(cln.row)}'] = ''
                if len(ln) > 1:
                    count = 0
                    for sl in ws[f'{pln_1}':f'{pln_2}']:
                        for cl in sl:
                            try:
                                ws[f'{letter[cl.column - 1] + str(cl.row)}'] = ln[count]
                                count += 1
                            except IndexError:
                                break
    return True


def LesPlan(is_class, is_next_day):
    les_plan = ['']
    for sheet_days in ws['B2':'B38']:
        for cell_days in sheet_days:
            if cell_days.value == f'{is_next_day}':
                position_class_1 = letter[cell_days.column + 1] + str(cell_days.row)
                position_class_2 = letter[cell_days.column + 11] + str(cell_days.row)
                for sheet_class in ws[f'{position_class_1}':f'{position_class_2}']:
                    for cell_class in sheet_class:
                        if cell_class.value == f'{is_class}':
                            position_plan_1 = letter[cell_class.column - 1] + str(cell_class.row + 1)
                            position_plan_2 = letter[cell_class.column - 1] + str(cell_class.row + 8)
                            for sheet_lesson in ws[f'{position_plan_1}':f'{position_plan_2}']:
                                for cell_lesson in sheet_lesson:
                                    if cell_lesson.value is None:
                                        break
                                    else:
                                        les_plan.append(cell_lesson.value)
                            result = {f'{is_class}': les_plan}
                            df = pandas.DataFrame(result)
                            return df
                    return 'Такого класса не существует'








# def OfficeNumber(is_continue):
#     spisok = set()
#     if is_continue:
#         for sheet_number in ws['D3':'N46']:
#             for cell_number in sheet_number:
#                 chislo = re.findall(r'\d{3}', str(cell_number.value))
#                 spisok.add(str(chislo))
#     print(*spisok)
#
#
# print(OfficeNumber(True))


# def les_plan(is_class):
#     book = openpyxl.open('C:\\Users\\Дмитрий\\PycharmProjects\\Da\\documents\\Расписание уроков.xlsx')
#     wsheet = book.active
#     for sheet_1 in wsheet['B2':'G9']:
#         for cell_2 in sheet_1:
#             if cell_2.value == f'{is_class}':
#                 posizia_1 = letter[cell_2.column - 1] + str(cell_2.row + 1)
#                 posizia_2 = letter[cell_2.column] + str(cell_2.row + 7)
#                 for cell_3 in wsheet[f'{posizia_1}':f'{posizia_2}']:
#                     for bruh in cell_3:
#                         if any(number.isdigit() for number in f'{bruh.value}'):
#                             stolb_2.append(bruh.value)
#                         elif bruh.value is None:
#                             break
#                         else:
#                             stolb_1.append(bruh.value)
#
#     vivod = {f'{is_class}': stolb_1,
#              'каб': stolb_2}
#     df = pandas.DataFrame(vivod)
#     df_1 = df
#     return df_1


# for row_1 in wsheet.iter_rows(min_row=3, max_col=3, max_row=9):
#     for cell_1 in row_1:
#         if cell_1.value == None:
#             pass
#         else:
#             print(cell_1.value)
# print('================')
# for sheet in wsheet['B3':'C8']:
#     for bruh in sheet:
#         if any(number.isdigit() for number in f'{bruh.value}'):
#             print(bruh.value)
#         else:
#             print(bruh.value, end=' ')
# print('================')


# stolb_1 = []
# stolb_2 = []
#
# cleal = input('Введи: ')
# book = openpyxl.open('C:\\Users\\Дмитрий\\PycharmProjects\\Da\\documents\\неа.xlsx')
# wsheet = book.active
# for sheet_1 in wsheet['B2':'G9']:
#     for cell_2 in sheet_1:
#         if cell_2.value == f'{cleal}':
#             posizia_1 = letter[cell_2.column - 1] + str(cell_2.row + 1)
#             posizia_2 = letter[cell_2.column] + str(cell_2.row + 7)
#             for cell_3 in wsheet[f'{posizia_1}':f'{posizia_2}']:
#                 for bruh in cell_3:
#                     if any(number.isdigit() for number in f'{bruh.value}'):
#                         stolb_2.append(bruh.value)
#                     elif bruh.value is None:
#                         break
#                     else:
#                         stolb_1.append(bruh.value)
